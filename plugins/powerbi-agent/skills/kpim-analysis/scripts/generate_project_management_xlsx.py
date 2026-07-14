import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
HDR = PatternFill("solid", fgColor="3F5B8C"); HF = Font(bold=True, color="FFFFFF", size=11)
SUB = PatternFill("solid", fgColor="0F9D8F"); SF = Font(bold=True, color="FFFFFF")
PH = PatternFill("solid", fgColor="DDE5F2")
thin = Side(style="thin", color="C9D3E6"); BD = Border(thin,thin,thin,thin)
wrap = Alignment(wrap_text=True, vertical="top")
def style_header(ws, row, ncol):
    for c in range(1,ncol+1):
        cell=ws.cell(row=row,column=c); cell.fill=HDR; cell.font=HF; cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=BD
def put(ws, r, vals, border=True):
    for i,v in enumerate(vals,1):
        cell=ws.cell(row=r,column=i,value=v); cell.alignment=wrap
        if border: cell.border=BD
def widths(ws, ws_widths):
    for i,w in enumerate(ws_widths,1): ws.column_dimensions[get_column_letter(i)].width=w

# ============ 1. KEY INFORMATION ============
ws=wb.active; ws.title="KEY INFORMATION"
ws["A1"]="KEY INFORMATION — KPIM Mart (Báo cáo Bán Hàng)"; ws["A1"].font=Font(bold=True,size=14,color="12776E")
r=3
sections=[
 ("1. REQUIREMENTS & OBJECTIVES",[
   ("Tổng hợp dữ liệu","Gộp SQL Server + Excel + SharePoint; kết nối bảng; so sánh TH vs KH; gộp nhiều bảng"),
   ("Tính toán thông minh","Chỉ số theo logic nghiệp vụ; time-intelligence (cùng kỳ, lũy kế); so sánh công thức đa bảng"),
   ("Trực quan hóa","Dashboard biểu đồ; Pivot Table; slicer & lọc điều kiện"),
   ("Cập nhật & Mở rộng","Kéo-thả phân tích mới; truy vết; dễ cập nhật; tương tác biểu đồ"),
 ]),
 ("2. ANALYTICS QUESTIONS",[
   ("Hiện trạng chỉ số bán hàng?","DT/số SP hiện tại; theo năm/tháng; theo cửa hàng/quản lý/sản phẩm"),
   ("Kết quả KD có đạt chỉ tiêu?","Cửa hàng/Quản lý/Tháng nào đạt hoặc không đạt chỉ tiêu"),
   ("Đang tăng trưởng hay suy giảm?","Tổng DT vs cùng kỳ; cửa hàng/SP tăng-giảm; tháng vs tháng trước, quý vs quý trước"),
   ("Yếu tố nào gây suy giảm DT?","Cửa hàng không đạt; vượt ngưỡng lũy kế; tháng giảm-cửa hàng giảm nhất; truy vết SP bán giảm"),
 ]),
 ("3. DATA REQUIRED",[
   ("SQL Server","Danh mục SP/KH/khu vực (Dimension)"),
   ("SharePoint List","Đơn hàng bán — Fact 15 trường"),
   ("Excel","Kế hoạch tháng theo khu vực (mỗi sheet 1 năm)"),
 ]),
 ("4. METRICS & DIMENSIONS",[
   ("Chỉ số (6 nhóm)","Doanh Thu; Lợi Nhuận Gộp; Tổng Khuyến Mãi; Số SP Bán; Số Khách Mua Hàng; Số Đơn Hàng"),
   ("Chiều (4 nhóm)","Thời gian; Khách hàng; Sản phẩm; Khu vực bán hàng"),
 ]),
 ("5. RESULT & DELIVERY",[
   ("6 báo cáo","Tổng quan; Phân tích doanh thu; Tái mua hàng; Kết quả theo khu vực; Phân khúc KH; Giám sát biên LN gộp"),
   ("Bàn giao",".pbix/.pbip + Data Model + measure + tài liệu md + Project_Management.xlsx + đào tạo"),
 ]),
]
put(ws,r,["Thành phần","Nội dung"]); style_header(ws,r,2); r+=1
for title,items in sections:
    ws.cell(row=r,column=1,value=title).fill=SUB; ws.cell(row=r,column=1).font=SF
    ws.cell(row=r,column=2).fill=SUB; ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2); r+=1
    for a,b in items:
        put(ws,r,[a,b]); r+=1
widths(ws,[34,90])

# ============ 2. PLANNING ============
ws=wb.create_sheet("PLANNING")
ws["A1"]="PLANNING — Kế hoạch triển khai (task 2 cấp)"; ws["A1"].font=Font(bold=True,size=14,color="12776E")
hdr=["Giai Đoạn","Đầu Mục Công Việc","Kết quả / Output","Tuần dự kiến","Trạng thái"]
put(ws,3,hdr); style_header(ws,3,len(hdr)); r=4
plan=[
 ("GĐ1. Khảo sát đánh giá",[
   ("1.1 Khảo sát yêu cầu","PROJECT.md (Key Information)","1",""),
   ("1.2 Khảo sát dữ liệu","RESEARCH_NOTES.md","1",""),
   ("1.3 Khảo sát nghiệp vụ","","1",""),
   ("1.4 Xác nhận nội dung thực hiện","User duyệt Key Information","2",""),
 ]),
 ("GĐ2. Xác nhận nguồn & kiến trúc Data Model",[
   ("2.1 Xác nhận nguồn dữ liệu đầu vào","Thông tin kết nối + data dictionary nguồn thật","2",""),
   ("2.2 Kiến trúc Data Model","Sơ đồ star schema","3",""),
   ("2.3 Đề xuất chỉnh sửa dữ liệu","Danh sách chỉnh sửa","3",""),
 ]),
 ("GĐ3. Kết nối, làm sạch & tải vào Data Model",[
   ("3.1 Kết nối Power Query (từng nguồn)","","3",""),
   ("3.2 Làm sạch (M Language)","","4",""),
   ("3.3 Load vào Power BI","","4",""),
   ("3.4 Kiểm tra quan hệ & chuẩn hóa Data Model","DATA_DICTIONARY.md","4",""),
   ("3.5 Chuẩn hóa Date Table & đặc tả model","","5",""),
   ("3.6 UAT bằng bảng Matrix","","5",""),
 ]),
 ("GĐ4. Tạo công thức DAX Measure",[
   ("4.1 Tạo bảng Measure","","5",""),
   ("4.2 Đối chiếu Analysis Mindmap + dữ liệu","METRICS_CALCULATION.md","6",""),
   ("4.3 Tạo folder con theo bộ hàm tính","","6",""),
   ("4.4 Tạo measure theo tên chuẩn","DOMAIN_DIMENSION.md","6",""),
   ("4.5 Update md + Excel","","7",""),
 ]),
 ("GĐ5. Tạo & thiết kế báo cáo",[
   ("5.1 Chốt danh sách báo cáo + xin asset","REPORTS.md","7",""),
   ("5.2 DESIGN.md + theme.json","theme.json","8",""),
   ("5.3 Xin mẫu báo cáo PBI + update theme","","8",""),
   ("5.4 Dựng từng báo cáo (Group→Report→Page)",".pbix/.pbip","9-12",""),
 ]),
]
for g,items in plan:
    ws.cell(row=r,column=1,value=g).fill=SUB; 
    for c in range(1,6): ws.cell(row=r,column=c).fill=SUB
    ws.cell(row=r,column=1).font=SF; r+=1
    for a,b,c,d in items:
        put(ws,r,["",a,b,c,d]); r+=1
widths(ws,[6,44,34,12,14])

# ============ 3. DATA DICTIONARY ============
ws=wb.create_sheet("DATA DICTIONARY")
ws["A1"]="DATA DICTIONARY — Fact Đơn Hàng Bán (15 trường)"; ws["A1"].font=Font(bold=True,size=14,color="12776E")
put(ws,3,["STT","Tên cột","Loại dữ liệu","Định nghĩa","Additive"]); style_header(ws,3,5); r=4
dd=[
(1,"Mã giao dịch","Text","Mã định danh giao dịch mua bán đơn hàng","Key"),
(2,"Ngày đặt hàng","Date","Ngày phát sinh đơn hàng, đặt bởi khách hàng","Key"),
(3,"Tên cửa hàng","List","Cửa hàng trên hệ thống phát sinh đơn hàng","Dim"),
(4,"Sản phẩm","List","Sản phẩm được đặt hàng","Dim"),
(5,"Khuyến mãi","List","Chương trình giảm giá áp dụng","Dim"),
(6,"Số lượng","Whole Number","Số lượng sản phẩm mua","Additive"),
(7,"Số điện thoại","Text","SĐT định danh khách hàng (PII)","Key/PII"),
(8,"Tên khách hàng","Text","Tên của khách hàng trên hệ thống","Attr"),
(9,"Giá bán","Decimal","Giá bán cho 1 đơn vị sản phẩm","Non-additive"),
(10,"Giá mua","Decimal","Giá vốn mua nhập sản phẩm về","Non-additive"),
(11,"Tổng giá tiền","Decimal","Số lượng × Giá bán","Additive"),
(12,"Tổng giảm giá","Decimal","Số lượng × Tỷ lệ giảm giá","Additive"),
(13,"Tổng doanh thu","Decimal","Tổng giá tiền − Tổng giảm giá","Additive"),
(14,"Giá vốn hàng bán","Decimal","Số lượng × Giá mua","Additive"),
(15,"Lợi nhuận gộp","Decimal","Tổng doanh thu − Giá vốn hàng bán","Additive"),
]
for row in dd:
    put(ws,r,list(row));
    if r%2==0:
        for c in range(1,6): ws.cell(row=r,column=c).fill=PH
    r+=1
widths(ws,[6,20,16,60,14])

# ============ 4. METRICS_CALCULATION ============
ws=wb.create_sheet("METRICS_CALCULATION")
ws["A1"]="METRICS_CALCULATION — DAX Measures"; ws["A1"].font=Font(bold=True,size=14,color="12776E")
put(ws,3,["Nhóm (Folder)","Measure","Mô tả","DAX gợi ý","Format"]); style_header(ws,3,5); r=4
mt=[
("Doanh Thu","Doanh Thu","Tổng doanh thu","SUM(Fact_DonHang[Tổng doanh thu])","#,0"),
("Doanh Thu","Doanh Thu LK Năm","Lũy kế năm (YTD)","TOTALYTD([Doanh Thu],Dim_Date[Date])","#,0"),
("Doanh Thu","Tăng trưởng vs Tháng trước","% vs tháng trước","DIVIDE([Doanh Thu]-CALCULATE([Doanh Thu],PREVIOUSMONTH(Dim_Date[Date])),...)","0.0%"),
("Doanh Thu","Tăng trưởng vs Cùng kỳ","% YoY","SAMEPERIODLASTYEAR → DIVIDE(TH-LY,LY)","0.0%"),
("Lợi Nhuận Gộp","Lợi Nhuận Gộp","Tổng LNG","SUM(Fact_DonHang[Lợi nhuận gộp])","#,0"),
("Lợi Nhuận Gộp","Giá Vốn Hàng Bán","Tổng giá vốn","SUM(Fact_DonHang[Giá vốn hàng bán])","#,0"),
("Lợi Nhuận Gộp","Biên Lợi Nhuận Gộp","LNG/Doanh thu","DIVIDE([Lợi Nhuận Gộp],[Doanh Thu])","0.0%"),
("Khuyến Mãi","Tổng Khuyến Mãi","Tổng giảm giá","SUM(Fact_DonHang[Tổng giảm giá])","#,0"),
("Khuyến Mãi","Tỷ Lệ Khuyến Mãi","Giảm giá/Tổng giá tiền","DIVIDE([Tổng Khuyến Mãi],SUM(Fact_DonHang[Tổng giá tiền]))","0.0%"),
("Khuyến Mãi","Số SP Chạy Khuyến Mãi","SP có KM","CALCULATE(DISTINCTCOUNT(Sản phẩm),Khuyến mãi<>\"Không\")","#,0"),
("Sản Phẩm","Số Sản Phẩm Bán","Tổng số lượng","SUM(Fact_DonHang[Số lượng])","#,0"),
("Sản Phẩm","Lợi Nhuận Trên 1 SP","LNG/số lượng","DIVIDE([Lợi Nhuận Gộp],[Số Sản Phẩm Bán])","#,0"),
("Sản Phẩm","Giá Bán TB 1 SP","DT/số lượng","DIVIDE([Doanh Thu],[Số Sản Phẩm Bán])","#,0"),
("Khách Hàng","Số Khách Mua Hàng","Khách distinct","DISTINCTCOUNT(Fact_DonHang[Số điện thoại])","#,0"),
("Khách Hàng","Tần Suất Mua TB 1 Khách","Số đơn/số khách","DIVIDE([Số Đơn Hàng],[Số Khách Mua Hàng])","#,0.0"),
("Khách Hàng","Bình Quân Giá Trị Mua 1 Khách","DT/số khách","DIVIDE([Doanh Thu],[Số Khách Mua Hàng])","#,0"),
("Đơn Hàng","Số Đơn Hàng","Giao dịch distinct","DISTINCTCOUNT(Fact_DonHang[Mã giao dịch])","#,0"),
("Đơn Hàng","Giá Trị TB 1 Đơn","DT/số đơn","DIVIDE([Doanh Thu],[Số Đơn Hàng])","#,0"),
("Đơn Hàng","Số SP Mua TB 1 Đơn","SL/số đơn","DIVIDE([Số Sản Phẩm Bán],[Số Đơn Hàng])","#,0.0"),
("Kế Hoạch","Doanh Thu Kế Hoạch","Tổng KH","SUM(Fact_KeHoach[Doanh thu KH])","#,0"),
("Kế Hoạch","% Hoàn Thành KH","TH/KH","DIVIDE([Doanh Thu],[Doanh Thu Kế Hoạch])","0.0%"),
("Kế Hoạch","Chênh Lệch vs KH","TH-KH","[Doanh Thu]-[Doanh Thu Kế Hoạch]","#,0"),
]
for row in mt:
    put(ws,r,list(row))
    if r%2==0:
        for c in range(1,6): ws.cell(row=r,column=c).fill=PH
    r+=1
widths(ws,[16,30,26,52,10])

# ============ 5. DOMAIN_DIMENSION ============
ws=wb.create_sheet("DOMAIN_DIMENSION")
ws["A1"]="DOMAIN_DIMENSION — Chiều phân tích & tư duy nghiệp vụ"; ws["A1"].font=Font(bold=True,size=14,color="12776E")
put(ws,3,["Nhóm chiều","Bảng","Cột chiều","Tư duy nghiệp vụ"]); style_header(ws,3,4); r=4
dm=[
("Thời gian","Dim_Date","Ngày/Tháng/Quý/Năm","Xu hướng; nền time-intelligence (YTD, cùng kỳ, tháng trước)"),
("Khách hàng","Dim_KhachHang","Phân khúc","Nhóm quan trọng/thường xuyên/ít mua → chăm sóc & KM"),
("Khách hàng","Dim_KhachHang","Hạng thẻ","Loyalty tier → ưu đãi & giữ chân"),
("Khách hàng","Dim_KhachHang","Nhóm tuổi","Chân dung KH → chọn SP/kênh"),
("Khách hàng","Dim_KhachHang","Giới tính","Hành vi mua theo giới"),
("Sản phẩm","Dim_SanPham","Ngành hàng","Cơ cấu DT theo ngành → đầu tư"),
("Sản phẩm","Dim_SanPham","Nhóm sản phẩm","Hiệu quả nhóm; cross-sell"),
("Sản phẩm","Dim_SanPham","Phân loại","So biên lợi nhuận theo loại"),
("Sản phẩm","Dim_SanPham","Nguồn gốc","Đánh giá xuất xứ/nhà cung cấp"),
("Sản phẩm","Dim_SanPham","Hãng, thương hiệu","Hiệu quả brand; rebate"),
("Khu vực","Dim_KhuVuc","Quận","Phân bố địa lý (bản đồ); độ phủ"),
("Khu vực","Dim_KhuVuc","Cửa hàng","Đơn vị đánh giá hiệu suất & chỉ tiêu"),
("Khu vực","Dim_KhuVuc","Quản lý","KPI & RLS theo quản lý"),
("Kịch bản","(scenario)","Thực hiện/Kế hoạch/Chỉ tiêu","So sánh % hoàn thành"),
]
for row in dm:
    put(ws,r,list(row))
    if r%2==0:
        for c in range(1,5): ws.cell(row=r,column=c).fill=PH
    r+=1
widths(ws,[14,16,22,60])

# ============ 6. REPORT ============
ws=wb.create_sheet("REPORT")
ws["A1"]="REPORT — Danh sách & đặc tả báo cáo (Group→Report→Page)"; ws["A1"].font=Font(bold=True,size=14,color="12776E")
put(ws,3,["Report Group","Report","Trang","Mục tiêu","Visual (block kit)","Nguồn"]); style_header(ws,3,6); r=4
rp=[
("KPIM Mart — Bán Hàng","R1 Tổng Quan","1. Tổng quan chỉ số","DT/LNG/SP/đơn tổng","cardVisual, comboChart, slicer","SharePoint+SQL"),
("","","2. So sánh Quận/Cửa hàng","So chỉ số theo địa lý","azureMap, clusteredBarChart, pivotTable",""),
("","","3. So sánh Tháng/Năm & Ngành/SP","So theo thời gian & sản phẩm","comboChart, donutChart, pivotTable",""),
("","R2 Phân Tích Doanh Thu","1. DT vs Kế hoạch & Tăng trưởng","Doanh số/KH/%YoY","cardVisual, comboChart, slicer","SharePoint+Excel"),
("","","2. Xu hướng 12 tháng (nay vs năm ngoái)","So xu hướng","comboChart, cardVisual",""),
("","","3. DT theo Quận/Cửa hàng + Lũy kế vs Chỉ tiêu","Giám sát tăng trưởng & lũy kế","pivotTable, clusteredBarChart, azureMap",""),
("","R3 Phân Tích Tái Mua Hàng","1. KH mới/cũ/duy trì vs tháng trước","Cấu trúc khách theo trạng thái","cardVisual, comboChart, donutChart","SharePoint"),
("","","2. Tỷ trọng KH quay lại (quý/nửa năm/năm)","Retention/cohort","clusteredBarChart, pivotTable",""),
("","R4 Kết Quả Theo Khu Vực","1. Cửa hàng đạt/không đạt vs năm ngoái","Xếp hạng đạt chỉ tiêu","cardVisual, clusteredBarChart, pivotTable","SharePoint+Excel"),
("","","2. % tiến độ DT đạt chỉ tiêu từng cửa hàng","Bullet tiến độ","clusteredBarChart, pivotTable",""),
("","","3. % tăng trưởng cùng kỳ từng cửa hàng","So YoY theo cửa hàng","azureMap, clusteredBarChart",""),
("","R5 Phân Khúc Khách Hàng","1. Giá trị đơn TB/khách","Chỉ số theo khách","cardVisual, scatterChart","SharePoint+SQL"),
("","","2. Số KH theo phân loại (RFM)","Segment","donutChart, clusteredBarChart, pivotTable",""),
("","","3. DT theo nhóm KH (tuổi/nghề/giới)","Chân dung khách","clusteredBarChart, pivotTable",""),
("","R6 Giám Sát Biên LN Gộp","1. Tổng quan DS/giá vốn/LNG/biên","KPI lợi nhuận","cardVisual, comboChart","SharePoint+SQL"),
("","","2. LNG theo từng cửa hàng","So sánh cửa hàng","clusteredBarChart, pivotTable",""),
("","","3. Giá bán/mua/LN & biên từng SP","Chi tiết sản phẩm","scatterChart, pivotTable",""),
]
for row in rp:
    put(ws,r,list(row))
    if r%2==0:
        for c in range(1,7): ws.cell(row=r,column=c).fill=PH
    r+=1
widths(ws,[22,24,34,30,34,16])

# freeze headers
for name in wb.sheetnames:
    wb[name].freeze_panes="A4"

out="./Project_Management.xlsx"
wb.save(out)
print("saved", out, "| sheets:", wb.sheetnames)
